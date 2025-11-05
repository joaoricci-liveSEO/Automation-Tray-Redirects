# Este é o automation_logic.py

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC 
import openpyxl
import time
import os
import sys

# --- Configurações de Seletores (Mantidas) ---
BTN_ADICIONAR_REDIRECIONAMENTO_XPATH = "//button[contains(@class, 'btn-primary') and contains(., 'Adicionar redirecionamento')]"
INPUT_URL_ORIGEM_ID = "redirection-input-source_url" 
INPUT_URL_DESTINO_ID = "redirection-input-target_url" 
BTN_SALVAR_MODAL_XPATH = "//button[contains(@class, 'redirection-url-sidebar__button--save') and contains(text(), 'Salvar')]" 

# --- Configurações Iniciais ---
# Esta função ajuda a encontrar o .xlsx depois que o app vira .exe
def resource_path(relative_path):
    """ Retorna o caminho absoluto para o recurso, funcionando em dev e no PyInstaller """
    try:
        # PyInstaller cria uma pasta temp e armazena o caminho em _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

NOME_ARQUIVO = resource_path('de_para_redirecionamento.xlsx')
COLUNA_STATUS = 3 
URL_REDIRECIONAMENTO_TRAY = "https://colorepedrarias.commercesuite.com.br/admin/settings/redirect?sort=-id&page%5Bsize%5D=25&page%5Bnumber%5D=1" 

# Objeto auxiliar para que nossos 'print()' sejam enviados para a GUI
class GuiLogger:
    def __init__(self, log_queue):
        self.log_queue = log_queue

    def write(self, message):
        if message.strip(): # Evita linhas em branco
            self.log_queue.put(message)

    def flush(self):
        pass # Necessário para a interface de 'file-like object'

def iniciar_automacao():
    """Inicia o WebDriver do Chrome."""
    service = Service(ChromeDriverManager().install())
    chrome_options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.maximize_window()
    return driver

# --- FUNÇÃO DE PAUSA MODIFICADA ---
def pausar_e_perguntar_salvamento(workbook, driver, motivo_erro, row_index, log_queue, decision_queue):
    """
    Pausa a execução, informa o usuário (via GUI) e espera uma decisão.
    """
    log_queue.put(f"--- 🚨 ATENÇÃO: AUTOMAÇÃO PAUSADA 🚨 ---")
    log_queue.put(f"Motivo: {motivo_erro}")
    if row_index:
        log_queue.put(f"O erro ocorreu ao processar a linha: {row_index}")

    # 1. Informa à GUI que ela deve mostrar o pop-up de decisão
    log_queue.put(("SHOW_PAUSE_DIALOG", motivo_erro))
    
    # 2. Bloqueia esta thread (de automação) até a GUI (thread principal) colocar uma decisão na fila
    log_queue.put("Aguardando decisão do usuário (S/T/F)...")
    user_choice = decision_queue.get() # Isso vai travar a thread_worker até a GUI responder
    log_queue.put(f"Usuário escolheu: {user_choice}")

    if user_choice == 'S':
        log_queue.put("Salvando progresso parcial...")
        try:
            workbook.save(NOME_ARQUIVO)
            log_queue.put(f"Planilha '{os.path.basename(NOME_ARQUIVO)}' salva com o progresso atual.")
        except Exception as save_e:
            log_queue.put(f"ERRO CRÍTICO ao tentar salvar a planilha: {save_e}")
        
        log_queue.put("Encerrando a automação.")
        if driver:
            driver.quit()
        return "EXIT" # Sinaliza para sair

    elif user_choice == 'T':
        if not driver:
            log_queue.put("O driver não está disponível. Não é possível tentar novamente.")
            return "EXIT"
            
        log_queue.put("Tentando recarregar a página para continuar...")
        try:
            driver.refresh()
            WebDriverWait(driver, 20).until(
                EC.visibility_of_element_located((By.XPATH, BTN_ADICIONAR_REDIRECIONAMENTO_XPATH))
            )
            log_queue.put("Página recarregada. A automação continuará na *PRÓXIMA* linha.")
            return "CONTINUE" # Sinaliza para continuar o loop
        except Exception as refresh_e:
            log_queue.put(f"Falha ao tentar recarregar a página: {refresh_e}")
            log_queue.put("Não foi possível recuperar. Encerrando.")
            return "EXIT" # Sinaliza para sair

    elif user_choice == 'F':
        log_queue.put("Saindo sem salvar.")
        if driver:
            driver.quit()
        return "EXIT" # Sinaliza para sair

# --- FUNÇÃO PRINCIPAL MODIFICADA ---
# NO ARQUIVO: automation_logic.py
# (Substitua a função inteira 'executar_automacao' por esta)

def executar_automacao(log_queue, decision_queue):
    """
    Função principal da automação, modificada para aceitar queues
    para logging (log_queue) e decisões (decision_queue).
    
    VERSÃO ATUALIZADA: Garante que o 'finally' sempre execute.
    """
    
    # Redireciona o 'print' (stdout) para a nossa fila da GUI
    sys.stdout = GuiLogger(log_queue)
    sys.stderr = GuiLogger(log_queue)

    driver = None # Definimos o driver aqui fora
    workbook = None # Definimos o workbook aqui fora
    
    # --- ESTE É O NOVO 'TRY' GIGANTE ---
    try:
        print("Iniciando processo de automação...")
        
        # --- Bloco 1: Carregar Planilha ---
        try:
            workbook = openpyxl.load_workbook(NOME_ARQUIVO)
            sheet = workbook.active
            print(f"Planilha '{os.path.basename(NOME_ARQUIVO)}' carregada.")
        except FileNotFoundError:
            print(f"ERRO: Arquivo '{os.path.basename(NOME_ARQUIVO)}' não encontrado.")
            return # O 'finally' lá embaixo VAI ser executado
        except Exception as e:
            print(f"ERRO ao carregar a planilha: {e}. Verifique se ela não está aberta.")
            return # O 'finally' lá embaixo VAI ser executado

        # --- Bloco 2: Iniciar Driver e Login ---
        try:
            print("Iniciando o navegador (ChromeDriver)...")
            driver = iniciar_automacao() # Atribui à variável 'driver'
            driver.get(URL_REDIRECIONAMENTO_TRAY)
            
            print(f"Aguarde... Você tem 20 segundos para fazer login, se necessário.")
            
            WebDriverWait(driver, 20).until(
                EC.visibility_of_element_located((By.XPATH, BTN_ADICIONAR_REDIRECIONAMENTO_XPATH))
            )
            print("Login detectado! Página de Redirecionamento carregada.")
            
        except Exception as e:
            # Este é o erro que você viu!
            print(f"ERRO: O botão 'Adicionar' não apareceu em 20 segundos.")
            print("Verifique o login ou a URL. Finalizando.")
            # Não precisamos do driver.quit() aqui, o 'finally' cuida disso
            return # O 'finally' lá embaixo VAI ser executado

        # --- Bloco 3: Loop Principal ---
        # (Este try/except interno é para o KeyboardInterrupt e erros globais *do loop*)
        try:
            for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2): 
                url_de = row[0].value
                url_para = row[1].value
                status = row[COLUNA_STATUS - 1].value 
                
                if status == True or not url_de or not url_para:
                    print(f"Linha {row_index}: Ignorada (Status TRUE ou URLs vazias).")
                    continue

                print(f"Processando Linha {row_index}: DE='{url_de}' PARA='{url_para}'")

                try:
                    btn_adicionar = driver.find_element(By.XPATH, BTN_ADICIONAR_REDIRECIONAMENTO_XPATH)
                    btn_adicionar.click()
                    time.sleep(1)

                    input_origem = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.ID, INPUT_URL_ORIGEM_ID))
                    )
                    input_origem.send_keys(url_de)

                    input_destino = driver.find_element(By.ID, INPUT_URL_ORIGEM_ID)
                    input_destino.send_keys(url_para)
                    
                    btn_salvar = driver.find_element(By.XPATH, BTN_SALVAR_MODAL_XPATH)
                    btn_salvar.click()
                    
                    time.sleep(4) 
                    
                    sheet.cell(row=row_index, column=COLUNA_STATUS, value=True)
                    print(f"Linha {row_index} processada e marcada como TRUE.")

                except Exception as e:
                    motivo = f"Erro ao processar linha {row_index} (possível CAPTCHA ou elemento não encontrado)."
                    action = pausar_e_perguntar_salvamento(workbook, driver, motivo, row_index, log_queue, decision_queue)
                    
                    if action == "EXIT":
                        print("Automação interrompida pelo usuário ou erro fatal.")
                        break # Sai do loop 'for'
                    elif action == "CONTINUE":
                        continue # Pula para a próxima linha do loop
            
            # Se o loop terminar sem 'break', salva
            else:
                print("\n✅ Processo de automação finalizado (loop concluído).")
                if workbook:
                    workbook.save(NOME_ARQUIVO)
                    print(f"Planilha '{os.path.basename(NOME_ARQUIVO)}' salva com sucesso.")

        except KeyboardInterrupt:
            motivo = "Interrupção manual (Ctrl+C) detectada."
            pausar_e_perguntar_salvamento(workbook, driver, motivo, None, log_queue, decision_queue)
        except Exception as e_global:
            motivo = f"Erro global inesperado na automação: {e_global}"
            pausar_e_perguntar_salvamento(workbook, driver, motivo, None, log_queue, decision_queue)

    # --- ESTE É O NOVO 'FINALLY' GIGANTE ---
    # Ele será executado SEMPRE, não importa onde o 'return' ou erro ocorra.
    finally:
        if driver and driver.session_id:
            try:
                driver.quit()
                print("Navegador fechado.")
            except: # Ignora erros ao fechar o driver (pode já estar fechado)
                pass
        
        # Esta é a mensagem que a GUI está esperando para resetar o botão!
        print("Thread de automação finalizada.")