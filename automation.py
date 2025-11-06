from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC 
import openpyxl
import time
import sys # Importado para o sys.exit()

# --- Configurações de Seletores (Mantidas) ---
BTN_ADICIONAR_REDIRECIONAMENTO_XPATH = "//button[contains(@class, 'btn-primary') and contains(., 'Adicionar redirecionamento')]"
INPUT_URL_ORIGEM_ID = "redirection-input-source_url" 
INPUT_URL_DESTINO_ID = "redirection-input-target_url" 
BTN_SALVAR_MODAL_XPATH = "//button[contains(@class, 'redirection-url-sidebar__button--save') and contains(text(), 'Salvar')]" 

# --- Configurações Iniciais (Mantidas) ---
NOME_ARQUIVO = 'de_para_redirecionamento.xlsx'
COLUNA_STATUS = 3 
URL_REDIRECIONAMENTO_TRAY = "https://colorepedrarias.commercesuite.com.br/admin/settings/redirect?sort=-id&page%5Bsize%5D=25&page%5Bnumber%5D=1" 

def iniciar_automacao():
    """Inicia o WebDriver do Chrome."""
    service = Service(ChromeDriverManager().install())
    chrome_options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.maximize_window()
    return driver

# --- NOVA FUNÇÃO AUXILIAR PARA PAUSA E SALVAMENTO ---
def pausar_e_perguntar_salvamento(workbook, driver, motivo_erro, row_index=None):
    """
    Pausa a execução, informa o usuário e pergunta se deseja salvar e sair.
    Esta função agora centraliza toda a lógica de pausa.
    """
    print(f"\n--- 🚨 ATENÇÃO: AUTOMAÇÃO PAUSADA 🚨 ---")
    print(f"Motivo: {motivo_erro}")
    if row_index:
        print(f"O erro ocorreu ao processar a linha: {row_index}")

    while True:
        user_choice = input("Você deseja:\n"
                            " (S) Salvar o progresso parcial (baixar planilha) e SAIR.\n"
                            " (T) Tentar recuperar (refresh na página) e CONTINUAR (ignora a linha atual se houver erro nela).\n"
                            " (F) FORÇAR SAÍDA sem salvar.\n"
                            "Escolha (S/T/F): ").strip().upper()

        if user_choice == 'S':
            # User quer salvar e sair.
            print("Salvando progresso parcial...")
            try:
                workbook.save(NOME_ARQUIVO)
                print(f"Planilha '{NOME_ARQUIVO}' salva com o progresso atual.")
            except Exception as save_e:
                print(f"ERRO CRÍTICO ao tentar salvar a planilha: {save_e}")
                print(f"O arquivo pode estar aberto ou sem permissão. Tente salvar manualmente.")
            
            print("Encerrando a automação.")
            if driver:
                driver.quit()
            sys.exit() # Força a saída do script

        elif user_choice == 'T':
            # User quer tentar recarregar e continuar.
            if not driver:
                print("O driver não está disponível. Não é possível tentar novamente.")
                continue # Volta ao loop de pergunta
                
            print("Tentando recarregar a página para continuar...")
            try:
                driver.refresh()
                # Espera a página estar pronta novamente
                WebDriverWait(driver, 20).until(
                    EC.visibility_of_element_located((By.XPATH, BTN_ADICIONAR_REDIRECIONAMENTO_XPATH))
                )
                print("Página recarregada. A automação continuará na *PRÓXIMA* linha.")
                return # Sai da função de pausa e permite que o loop continue
            except Exception as refresh_e:
                print(f"Falha ao tentar recarregar a página: {refresh_e}")
                print("Voltando ao menu de pausa. O erro persiste.")
                # O 'while True' da pausa continuará

        elif user_choice == 'F':
            print("Saindo sem salvar. O navegador pode permanecer aberto.")
            if driver:
                driver.quit()
            sys.exit() # Força a saída do script

        else:
            print("Opção inválida. Escolha 'S', 'T' ou 'F'.")


def executar_automacao():
    # Carregar a planilha
    try:
        workbook = openpyxl.load_workbook(NOME_ARQUIVO)
        sheet = workbook.active
    except FileNotFoundError:
        print(f"ERRO: Arquivo '{NOME_ARQUIVO}' não encontrado.")
        return
    except Exception as e:
        print(f"ERRO ao carregar a planilha '{NOME_ARQUIVO}'. Verifique se ela não está corrompida ou aberta por outro programa.")
        print(f"Detalhe: {e}")
        return

    driver = iniciar_automacao()
    driver.get(URL_REDIRECIONAMENTO_TRAY)
    
    print(f"Acessando a página de redirecionamento. Você tem 20 segundos para completar o login, se necessário.")
    
    try:
        WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((By.XPATH, BTN_ADICIONAR_REDIRECIONAMENTO_XPATH))
        )
        print("Página de Redirecionamento carregada. Iniciando os redirecionamentos...")
    except Exception as e:
        print(f"ERRO: O botão 'Adicionar redirecionamento' não apareceu em 20 segundos.")
        print("Verifique o login ou a URL. Finalizando o script.")
        driver.quit()
        return

    # --- NOVO BLOCO try/except/finally ADICIONADO ---
    # Este bloco envolve todo o processo para capturar interrupções (Ctrl+C)
    # e garantir que o driver seja fechado.
    try:
        # Itera pelas linhas da planilha (começando da linha 2, após o cabeçalho)
        for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2): 
            url_de = row[0].value
            url_para = row[1].value
            status = row[COLUNA_STATUS - 1].value 
            
            if status == True or not url_de or not url_para:
                print(f"Linha {row_index}: Ignorada (Status TRUE ou URLs vazias).")
                continue

            print(f"Processando Linha {row_index}: DE='{url_de}' PARA='{url_para}'")

            # --- BLOCO try/except ATUALIZADO ---
            # Este bloco agora chama a nova função de pausa em caso de erro.
            try:
                # 4 - Executar o clique em "+Adicionar redirecionamento"
                btn_adicionar = driver.find_element(By.XPATH, BTN_ADICIONAR_REDIRECIONAMENTO_XPATH)
                btn_adicionar.click()
                time.sleep(1) # Espera o modal abrir

                # 5 - Colar a URL DE (Usando o ID)
                input_origem = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.ID, INPUT_URL_ORIGEM_ID))
                )
                input_origem.send_keys(url_de)

                # 6 - Colar a URL PARA (Usando o ID)
                input_destino = driver.find_element(By.ID, INPUT_URL_DESTINO_ID)
                input_destino.send_keys(url_para)
                
                # 7 - Clicar no botão salvar
                btn_salvar = driver.find_element(By.XPATH, BTN_SALVAR_MODAL_XPATH)
                btn_salvar.click()
                
                # Espera otimizada (ex: esperar o modal de sucesso ou o modal fechar)
                # Por enquanto, mantemos o time.sleep(4)
                time.sleep(4) 
                
                # 8 - Marcar na planilha
                sheet.cell(row=row_index, column=COLUNA_STATUS, value=True)
                print(f"Linha {row_index} processada e marcada como TRUE.")

            except Exception as e:
                # Em vez de apenas dar refresh, chamamos a função de pausa
                motivo = f"Erro inesperado ao processar a linha (possível CAPTCHA ou elemento não encontrado). Detalhe: {e}"
                pausar_e_perguntar_salvamento(workbook, driver, motivo, row_index)
                # Se o usuário escolher 'T' (Tentar novamente), a função retorna aqui 
                # e o loop 'for' continua para a PRÓXIMA linha.
            # --- FIM DO BLOCO try/except ATUALIZADO ---

        # Se o loop terminar sem interrupção, salva o arquivo
        print("\n✅ Processo de automação finalizado (loop concluído).")
        workbook.save(NOME_ARQUIVO)
        print(f"Planilha '{NOME_ARQUIVO}' salva com sucesso.")

    except KeyboardInterrupt:
        # Se o usuário pressionar Ctrl+C
        motivo = "Interrupção manual (Ctrl+C) detectada."
        pausar_e_perguntar_salvamento(workbook, driver, motivo)

    except Exception as e_global:
        # Captura qualquer outro erro catastrófico fora do loop
        motivo = f"Erro global inesperado na automação: {e_global}"
        pausar_e_perguntar_salvamento(workbook, driver, motivo)
        
    finally:
        # Este bloco SEMPRE é executado, exceto se sys.exit() for chamado
        if driver and driver.session_id:
            driver.quit()
            print("Navegador fechado.")
        
if __name__ == "__main__":
    executar_automacao()